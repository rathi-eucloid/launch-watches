#!/usr/bin/env python3
"""
script for scraping product pages from Amazon, BestBuy, and Samsung using Playwright.
- Saves each page's HTML to a unique file in outputs/ directory.
- Parses each HTML to extract product price and model number (or SKU for Samsung).

This version additionally saves the run's result as a single row in an Excel file:
- Columns = keys (flattened per-site/per-index keys like "amazon_1_url")
- Row = values for this run
- On next run the script appends a new row (does not overwrite previous data).

"""
from zoneinfo import ZoneInfo
import datetime
import asyncio
import json
import os
import random
import re
import socket
import shutil
import tempfile
import subprocess
import urllib.request

from urllib.parse import quote_plus
from playwright.async_api import async_playwright, TimeoutError
from bs4 import BeautifulSoup

# Optional anti-bot stealth (used for BestBuy only). The package API differs
# across versions, so we detect what's available and expose a single async
# helper _apply_stealth(page). If the package isn't installed, it's a no-op so
# the rest of the script still runs.
#   - playwright-stealth 1.x:  from playwright_stealth import stealth_async
#   - playwright-stealth 2.x:  from playwright_stealth import Stealth  (Stealth().apply_stealth_async)
try:
    from playwright_stealth import stealth_async as _stealth_async  # 1.x

    async def _apply_stealth(page):
        await _stealth_async(page)
    _STEALTH_AVAILABLE = True
except Exception:
    try:
        from playwright_stealth import Stealth as _Stealth  # 2.x
        _stealth_instance = _Stealth()

        async def _apply_stealth(page):
            await _stealth_instance.apply_stealth_async(page)
        _STEALTH_AVAILABLE = True
    except Exception:
        async def _apply_stealth(page):
            return None  # package not installed -> no-op
        _STEALTH_AVAILABLE = False
from openpyxl.utils import column_index_from_string, get_column_letter
# new imports for Excel writing
from openpyxl import Workbook, load_workbook

# =========================================================================
# script_2.py fixes (vs script.py)
# -------------------------------------------------------------------------
# 1. Redirect detection: many product pages now silently redirect to a
#    *different* product (e.g. S26 Ultra -> S25 FE, Z Fold 7 -> Z Flip 7) when
#    the item is out of stock/unavailable. We compare the product identifier we
#    REQUESTED (ASIN / BestBuy code / Samsung SKU) against the identifier in the
#    page's <link rel="canonical">. On mismatch we record "not available".
# 2. Duplicate-element fix: the old hardcoded selectors matched many elements on
#    the page (related items, sponsored, carousels), causing wrong reads. We now
#    scope price extraction to the MAIN price container only.
# 3. Updated selectors for current UI: Amazon price -> corePriceDisplay /
#    priceToPay; BestBuy & Samsung price -> JSON-LD offer (Samsung keyed by SKU).
# 4. BestBuy fix: headless Chromium fails with ERR_HTTP2_PROTOCOL_ERROR (BestBuy
#    tears down the HTTP/2 connection for headless browsers), so nothing is saved.
#    We launch BestBuy's Chromium HEADFUL (headless=False) on default HTTP/2 (NOT
#    --disable-http2, which made the CDN return empty shells) and with no custom
#    user-agent. On a display-less server (EC2 / GitHub Actions) run it under Xvfb.
# 5. results.xlsx now uses the same layout as "Price Comparisons_v3_WIP":
#    51 product groups x 9 columns starting at column C, timestamp in column B.
# Everything else (URLs, delays, user agents, cookies logic) is unchanged.
# =========================================================================
NOT_AVAILABLE = "not available"

# Retry policy for transient failures (network errors, timeouts, or a page that
# navigated OK but didn't render its price in time). 1 initial try + 2 retries.
# Genuine outcomes (a redirect to another product, or a page that explicitly says
# it's unavailable) are treated as FINAL and are NOT retried.
MAX_ATTEMPTS = 3
RETRY_BACKOFF_SEC = 5

# Excel layout of Price Comparisons_v3_WIP (per product group of 9 columns):
#   +0 Amazon price  +1 Samsung price  +2 BestBuy price
#   +3 SKU_Amazon    +4 SKU_Samsung    +5 SKU_BestBuy
#   +6 vs Amazon (formula, left blank)  +7 vs Bestbuy (formula, left blank)  +8 blank
FIRST_GROUP_COL = 3        # column C
GROUP_STRIDE    = 9
TIMESTAMP_COL   = 2        # column B

# Product group headers for row 1 (matches the WIP workbook, 51 slots in order)
PRODUCT_LABELS = [
    "Galaxy Watch9 40mm Cream Bluetooth",
    "Galaxy Watch9 40mm Cream LTE",

    "Galaxy Watch9 40mm Graphite Bluetooth",
    "Galaxy Watch9 40mm Graphite LTE",

    "Galaxy Watch9 44mm Silver Bluetooth",
    "Galaxy Watch9 44mm Silver LTE",

    "Galaxy Watch9 44mm Graphite Bluetooth",
    "Galaxy Watch9 44mm Graphite LTE",

    "Galaxy Watch Ultra2 47mm Titanium Gray LTE",
    "Galaxy Watch Ultra2 47mm Titanium Silver LTE"
]

SUBHEADERS = ["Amazon price", "Samsung price", "BestBuy.com price",
              "SKU_ID_Amazon ", "SKU_ID_Samsung ", "SKU_ID_BestBuy.com",
              "vs Amazon", "vs Bestbuy"]


# ---- product identifiers (used for redirect detection & slot SKUs) ----
def amazon_id_from_url(url):
    m = re.search(r"/dp/([A-Z0-9]{10})", url or "", re.I)
    return m.group(1).upper() if m else None

def bestbuy_id_from_url(url):
    m = re.search(r"/product/[^/]+/([A-Z0-9]+)", url or "", re.I)
    return m.group(1).upper() if m else None

def get_canonical_href(html):
    """Pull <link rel=canonical href=...> without a full DOM parse."""
    m = re.search(r'<link\b[^>]*\brel=["\']canonical["\'][^>]*>', html, re.I)
    if not m:
        return None
    h = re.search(r'href=["\']([^"\']+)["\']', m.group(0), re.I)
    return h.group(1) if h else None

def _looks_unavailable(html, site):
    """True when the page EXPLICITLY signals the product is unavailable/sold out.

    Used by the retry loop to decide whether a "no price" outcome is FINAL (the
    seller genuinely isn't selling it -> don't waste retries) vs TRANSIENT (the
    price element simply didn't render this time -> retry). Verified against the
    saved pages: Amazon's own out-of-stock listings render the exact phrase
    "Currently unavailable"; only third-party offers carry a price, which we
    deliberately don't scrape.
    """
    if not html:
        return False
    low = html.lower()
    if site == "amazon":
        return "currently unavailable" in low
    if site == "bestbuy":
        return "sold out" in low or "no longer available" in low
    if site == "samsung":
        return ("sold out" in low or "coming soon" in low
                or "out of stock" in low or "notify me" in low)
    return False


def _amazon_buybox_is_used(html):
    """True when the WINNING Amazon buybox offer is a USED/renewed device.

    We only want NEW-device prices. Some listings (e.g. a couple of S25 Edge
    variants) have a USED offer as the featured buybox, so the main price
    container shows the used price. We must NOT capture that.

    Two precise signals, verified against the saved pages:
      1. <div id="usedBuySection"> — Amazon renders this only when the featured
         buybox offer's condition is used ("Buy used: $...").
      2. A "Used: <condition>" label in the buybox (Like New / Very Good / Good /
         Acceptable).
    Both fire together on used-buybox pages and on NONE of the new-condition
    pages — including listings that merely OFFER a used alternative in a separate
    accordion (their buybox winner is still new), so this does not false-positive.
    """
    if not html:
        return False
    if re.search(r'id=["\']usedBuySection["\']', html):
        return True
    if re.search(r'Used:\s*(Like New|Very Good|Good|Acceptable)', html, re.I):
        return True
    return False


def iter_ldjson(html):
    """Yield parsed JSON-LD objects from the HTML (regex-sliced, fast)."""
    for m in re.finditer(
            r'<script\b[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html, re.I | re.S):
        try:
            data = json.loads(m.group(1).strip())
        except Exception:
            continue
        for it in (data if isinstance(data, list) else [data]):
            yield it


# ---- price cleaning (ported from ConvertDirtyTextPriceToNumbers/main2) ----
def clean_price_value(raw):
    """Return a float rounded to 2dp, or None. Mirrors main2_decimalPlaceTill2."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "" or s == NOT_AVAILABLE:
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    s = re.sub(r"[^\d\.,\-]", "", s)
    if s == "" or re.fullmatch(r"[-\.,]*", s):
        return None
    s = s.replace("−", "-")
    if "-" in s:
        if s.count("-") > 1:
            s = s.replace("-", "")
        if s.startswith("-"):
            negative = not negative
            s = s.lstrip("-")
    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
            if s.count(".") > 1:
                left, right = s.rsplit(".", 1)
                s = left.replace(".", "") + "." + right
    elif has_comma and not has_dot:
        parts = s.split(",")
        if len(parts) >= 2 and len(parts[-1]) == 2:
            s = ",".join(parts[:-1]).replace(",", "") + "." + parts[-1]
        else:
            s = s.replace(",", "")
    elif has_dot and not has_comma:
        if s.count(".") > 1:
            left, right = s.rsplit(".", 1)
            s = left.replace(".", "") + "." + right
    s = re.sub(r"[^\d.]", "", s)
    if s.count(".") > 1:
        left, right = s.rsplit(".", 1)
        s = left.replace(".", "") + "." + right
    if s in ("", "."):
        return None
    try:
        value = round(float(s), 2)
    except Exception:
        return None
    return -value if negative else value


# -----------------------
# Shared helpers
# -----------------------
async def human_delay(min_sec=0.5, max_sec=2.5):
    """Wait for a random time between min_sec and max_sec seconds."""
    delay = random.uniform(min_sec, max_sec)
    await asyncio.sleep(delay)

async def human_delay_short():
    """Small helper to yield control briefly (kept minimal to respect original logic)."""
    await asyncio.sleep(0.1)

async def get_page_content_safe(page, retries=4):
    """Return page HTML, tolerating in-flight client-side navigations.

    BestBuy fires a delayed client-side navigation/reload ~20s after load, which
    made a bare `page.content()` throw:
      "Unable to retrieve content because the page is navigating and changing".
    We wait for the page to settle and retry; as a last resort we read
    document.documentElement.outerHTML via JS (works mid-navigation).
    """
    last_err = None
    for attempt in range(retries):
        try:
            # let any in-flight navigation finish before grabbing content
            try:
                await page.wait_for_load_state("domcontentloaded", timeout=8000)
            except Exception:
                pass
            return await page.content()
        except Exception as e:
            last_err = e
            # brief settle, then retry
            await asyncio.sleep(2.5)
    # final fallback: pull the DOM directly (succeeds even while navigating)
    try:
        return await page.evaluate("() => document.documentElement.outerHTML")
    except Exception:
        raise last_err

def sanitize_filename(s: str, maxlen: int = 200) -> str:
    """Create a filesystem-safe short filename from a string (URL)."""
    if not s:
        return "file"
    s_enc = quote_plus(s, safe="")
    s_clean = re.sub(r'[^A-Za-z0-9._-]', '_', s_enc)
    return s_clean[:maxlen]

def _to_jsonable(v):
    """Convert complex types to JSON strings for Excel storage; leave primitives as-is."""
    if isinstance(v, (dict, list, tuple)):
        return json.dumps(v, ensure_ascii=False)
    return v

def _render_cells(result, slot_sku, site):
    """Given a per-URL result dict, return (price_cell, sku_cell) for the sheet.

    - price -> float when parseable, "not available" for redirect/no-price, else None.
    - sku   -> canonical per-slot SM code (Amazon/BestBuy upper, Samsung lower);
               "not available" mirrors the price when the product wasn't found.
    """
    raw = result.get("price") if result else None
    if raw == NOT_AVAILABLE:
        return NOT_AVAILABLE, NOT_AVAILABLE
    num = clean_price_value(raw)
    if num is None:
        return None, None            # genuine gap (fetch error / empty URL) -> blank
    sku = None
    if slot_sku:
        sku = slot_sku if site == "samsung" else slot_sku.upper()
    return num, sku


def save_results_wip_format(am_res, bb_res, sam_res, samsung_urls, ts_str,
                            excel_path="outputs/results.xlsx"):
    """Append one row per run to results.xlsx using the SAME layout as
    'Price Comparisons_v3_WIP.xlsx':
      - row 1 = product group headers, row 2 = sub-headers, data from row 3
      - 51 groups x 9 columns starting at column C; timestamp in column B
      - each group: Amazon/Samsung/BestBuy price, 3 SKU columns, 2 'vs'
        formula columns (filled with the same formulas as the WIP file), 1 blank
    Prices are written as numbers; SKU columns filled; 'vs' formulas added per
    row. Existing rows are kept.
    """
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=2, column=TIMESTAMP_COL, value="Timestamp EST")
        for s in range(len(PRODUCT_LABELS)):
            gc = FIRST_GROUP_COL + GROUP_STRIDE * s
            ws.cell(row=1, column=gc, value=PRODUCT_LABELS[s])
            for off, name in enumerate(SUBHEADERS):
                ws.cell(row=2, column=gc + off, value=name)

    r = ws.max_row + 1 if ws.max_row >= 2 else 3
    ws.cell(row=r, column=TIMESTAMP_COL, value=ts_str)

    n = len(PRODUCT_LABELS)
    for s in range(n):
        gc = FIRST_GROUP_COL + GROUP_STRIDE * s
        slot_sku = extract_sku_from_url(samsung_urls[s]) if s < len(samsung_urls) else None
        slot_sku = slot_sku.lower() if slot_sku else None

        for off, (res_list, site) in enumerate([
                (am_res, "amazon"), (sam_res, "samsung"), (bb_res, "bestbuy")]):
            result = res_list[s] if s < len(res_list) else None
            price_cell, sku_cell = _render_cells(result, slot_sku, site)
            pc = ws.cell(row=r, column=gc + off, value=price_cell)
            if isinstance(price_cell, float):
                pc.number_format = "0.00"
            ws.cell(row=r, column=gc + 3 + off, value=sku_cell)

        # +6 vs Amazon, +7 vs Bestbuy : same formulas as the WIP file, added on
        # every data row so they're in place as rows accumulate.
        #   vs Amazon  = Amazon price  / Samsung price - 1
        #   vs Bestbuy = BestBuy price / Samsung price - 1
        amazon_col  = get_column_letter(gc + 0)
        samsung_col = get_column_letter(gc + 1)
        bestbuy_col = get_column_letter(gc + 2)
        ws.cell(row=r, column=gc + 6,
                value=f"={amazon_col}{r}/{samsung_col}{r}-1")
        ws.cell(row=r, column=gc + 7,
                value=f"={bestbuy_col}{r}/{samsung_col}{r}-1")
        # +8 blank separator : intentionally left untouched

    wb.save(excel_path)
    print(f"✅ Results appended (WIP layout) to {excel_path} at row {r}")


def save_dict_to_excel_row(data: dict, excel_path: str = "outputs/results.xlsx"):
    """
    Save the provided dict as a single row in an Excel file.
    - Keys become column headers (first row).
    - Values become the next available row.
    - If the file exists, new keys are appended as new columns; existing column order is preserved.
    """
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        headers = list(data.keys())
        ws.append(headers)
        row = [ _to_jsonable(data.get(h)) for h in headers ]
        ws.append(row)
        wb.save(excel_path)
        print(f"✅ Results written to new Excel file: {excel_path}")
        return

    # file exists - load and append
    wb = load_workbook(excel_path)
    ws = wb.active

    # read existing headers from first row
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    existing_headers = [cell.value for cell in first_row]

    # compute headers union preserving existing order and appending new keys at the end
    new_keys = [k for k in data.keys() if k not in existing_headers]
    if new_keys:
        headers = existing_headers + new_keys
        # rewrite header row with expanded headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        headers = existing_headers

    # build row in header order
    row = []
    for h in headers:
        v = data.get(h)
        row.append(_to_jsonable(v) if v is not None else None)

    ws.append(row)
    wb.save(excel_path)

    #making changes from here
    column_refs = [
        "blank","a","d","cf","ar","e","cg","as","blank","blank","blank","i","ck","aw","j","cl","ax","blank","blank","blank","n","cp","bb","o","cq","bc","blank","blank","blank","s","cu","bg","t","cv","bh","blank","blank","blank","x","cz","bl","y","da","bm","blank","blank","blank","ac","de","bq","ad","df","br","blank","blank","blank","ah","dj","bv","ai","dk","bw","blank","blank","blank","am","do","ca","an","dp","cb","blank","blank","blank","dt","gb","ex","du","gc","ey","blank","blank","blank","dy","gg","fc","dz","gh","fd","blank","blank","blank","ed","gl","fh","ee","gm","fi","blank","blank","blank","ei","gq","fm","ej","gr","fn","blank","blank","blank","en","gv","fr","eo","gw","fs","blank","blank","blank","es","ha","fw","et","hb","fx", 'blank', 'blank', 'blank', 'hf', 'kr', 'iy', 'hg', 'ks', 'iz', 'blank', 'blank', 'blank', 'hk', 'kw', 'jd', 'hl', 'kx', 'je', 'blank', 'blank', 'blank', 'hp', 'lb', 'ji', 'hq', 'lc', 'jj', 'blank', 'blank', 'blank', 'hu', 'lg', 'jn', 'hv', 'lh', 'jo', 'blank', 'blank', 'blank', 'hz', 'll', 'js', 'ia', 'lm', 'jt', 'blank', 'blank', 'blank', 'ie', 'lq', 'jx', 'if', 'lr', 'jy', 'blank', 'blank', 'blank', 'ij', 'lv', 'kc', 'ik', 'lw', 'kd', 'blank', 'blank', 'blank', 'io', 'ma', 'kh', 'ip', 'mb', 'ki', 'blank', 'blank', 'blank', 'it', 'mf', 'km', 'iu', 'mg', 'kn', 'blank', 'blank', 'blank', 'mk', 'xe', 'ru', 'ml', 'xf', 'rv', 'blank', 'blank', 'blank', 'mp', 'xj', 'rz', 'mq', 'xk', 'sa', 'blank', 'blank', 'blank', 'mu', 'xo', 'se', 'mv', 'xp', 'sf', 'blank', 'blank', 'blank', 'mz', 'xt', 'sj', 'na', 'xu', 'sk', 'blank', 'blank', 'blank', 'ne', 'xy', 'so', 'nf', 'xz', 'sp', 'blank', 'blank', 'blank', 'nj', 'yd', 'st', 'nk', 'ye', 'su', 'blank', 'blank', 'blank', 'no', 'yi', 'sy', 'np', 'yj', 'sz', 'blank', 'blank', 'blank', 'nt', 'yn', 'td', 'nu', 'yo', 'te', 'blank', 'blank', 'blank', 'ny', 'ys', 'ti', 'nz', 'yt', 'tj', 'blank', 'blank', 'blank', 'od', 'yx', 'tn', 'oe', 'yy', 'to', 'blank', 'blank', 'blank', 'oi', 'zc', 'ts', 'oj', 'zd', 'tt', 'blank', 'blank', 'blank', 'on', 'zh', 'tx', 'oo', 'zi', 'ty', 'blank', 'blank', 'blank', 'os', 'zm', 'uc', 'ot', 'zn', 'ud', 'blank', 'blank', 'blank', 'ox', 'zr', 'uh', 'oy', 'zs', 'ui', 'blank', 'blank', 'blank', 'pc', 'zw', 'um', 'pd', 'zx', 'un', 'blank', 'blank', 'blank', 'ph', 'aab', 'ur', 'pi', 'aac', 'us', 'blank', 'blank', 'blank', 'pm', 'aag', 'uw', 'pn', 'aah', 'ux', 'blank', 'blank', 'blank', 'pr', 'aal', 'vb', 'ps', 'aam', 'vc', 'blank', 'blank', 'blank', 'pw', 'aaq', 'vg', 'px', 'aar', 'vh', 'blank', 'blank', 'blank', 'qb', 'aav', 'vl', 'qc', 'aaw', 'vm', 'blank', 'blank', 'blank', 'qg', 'aba', 'vq', 'qh', 'abb', 'vr', 'blank', 'blank', 'blank', 'ql', 'abf', 'vv', 'qm', 'abg', 'vw', 'blank', 'blank', 'blank', 'qq', 'abk', 'wa', 'qr', 'abl', 'wb', 'blank', 'blank', 'blank', 'qv', 'abp', 'wf', 'qw', 'abq', 'wg', 'blank', 'blank', 'blank', 'ra', 'abu', 'wk', 'rb', 'abv', 'wl', 'blank', 'blank', 'blank', 'rf', 'abz', 'wp', 'rg', 'aca', 'wq', 'blank', 'blank', 'blank', 'rk', 'ace', 'wu', 'rl', 'acf', 'wv', 'blank', 'blank', 'blank', 'rp', 'acj', 'wz', 'rq', 'ack', 'xa',
    ]
    new_sheet_base_name="SelectedColumns"
    source_sheet_name=None
    # select source sheet
    if source_sheet_name:
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{source_sheet_name}' not found in workbook.")
        src = wb[source_sheet_name]
    else:
        src = wb[wb.sheetnames[0]]

    # create unique new sheet name
    # new_name = new_sheet_base_name
    new_name = "converted"
    if new_name in wb.sheetnames:
        del wb["converted"]
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    tgt = wb.create_sheet(title=new_name)
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    # tgt = wb.create_sheet(title=new_name)

    max_row = src.max_row if src.max_row is not None else 0

    # target column pointer (1-indexed for openpyxl)
    tgt_col_idx = 1

    for token in column_refs:
        is_blank = token is None or (isinstance(token, str) and token.strip().lower() == "blank column")
        if is_blank:
            # leave a blank column (i.e., do nothing but advance tgt_col_idx)
            tgt_col_idx += 1
            continue

        # try to interpret token as Excel column letters
        col_letters = str(token).strip()
        try:
            src_col_idx = column_index_from_string(col_letters.upper())
        except Exception:
            # invalid column reference — create an empty column instead
            for r in range(1, max_row + 1):
                tgt.cell(row=r, column=tgt_col_idx, value=None)
            tgt_col_idx += 1
            continue

        # Copy values from source column to target column
        for r in range(1, max_row + 1):
            src_cell = src.cell(row=r, column=src_col_idx)
            # copy value only (not style/formula). If formula needed, assign src_cell.value (it will copy the formula text)
            tgt.cell(row=r, column=tgt_col_idx, value=src_cell.value)
        tgt_col_idx += 1
    wb.save(excel_path)
    print(f"✅ Results appended to Excel file: {excel_path}")





def copy_columns_by_references(
    file_path: str,
    column_refs: list,
    source_sheet_name: str | None = None,
    new_sheet_base_name: str = "CopiedColumns"
) -> str:
    """
    Copy columns from source sheet to a new sheet using Excel column letters.
    - column_refs: list of strings, column letters like ['A','D','X','AR', ...] or 'blank column'
    - source_sheet_name: None -> first sheet is used
    Returns the name of the created sheet.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path)
    # select source sheet
    if source_sheet_name:
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{source_sheet_name}' not found in workbook.")
        src = wb[source_sheet_name]
    else:
        src = wb[wb.sheetnames[0]]

    # create unique new sheet name
    new_name = "converted"
    if new_name in wb.sheetnames:
        del wb["converted"]
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    tgt = wb.create_sheet(title=new_name)

    max_row = src.max_row if src.max_row is not None else 0

    # target column pointer (1-indexed for openpyxl)
    tgt_col_idx = 1

    for token in column_refs:
        is_blank = token is None or (isinstance(token, str) and token.strip().lower() == "blank column")
        if is_blank:
            # leave a blank column (i.e., do nothing but advance tgt_col_idx)
            tgt_col_idx += 1
            continue

        # try to interpret token as Excel column letters
        col_letters = str(token).strip()
        try:
            src_col_idx = column_index_from_string(col_letters.upper())
        except Exception:
            # invalid column reference — create an empty column instead
            for r in range(1, max_row + 1):
                tgt.cell(row=r, column=tgt_col_idx, value=None)
            tgt_col_idx += 1
            continue

        # Copy values from source column to target column
        for r in range(1, max_row + 1):
            src_cell = src.cell(row=r, column=src_col_idx)
            # copy value only (not style/formula). If formula needed, assign src_cell.value (it will copy the formula text)
            tgt.cell(row=r, column=tgt_col_idx, value=src_cell.value)
        tgt_col_idx += 1

    # Save workbook (overwrites existing file)
    wb.save(file_path)
    return new_name


# -----------------------
# AMAZON-specific logic
# -----------------------
async def save_amazon_htmls(
    urls,
    output_dir="outputs",
    cookies_file="amazon_cookies.json",
    headless=True,
):
    """Loop over the list of URLs, save each HTML to a unique file, and update cookies once."""
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless, slow_mo=100)

        # Load existing cookies/session state if available
        if os.path.exists(cookies_file):
            print("🍪 Loading existing cookies/session...")
            context = await browser.new_context(storage_state=cookies_file)
        else:
            print("🆕 No cookies found, creating a new session...")
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1366, "height": 768},
            )

        try:
            results = []
            for idx, url in enumerate(urls, start=1):
                # empty slot (e.g. product not yet listed): keep the position so
                # results stay aligned with the product groups, but skip cleanly.
                if not url or not url.strip():
                    print(f"\n[Amazon {idx}/{len(urls)}] empty URL slot -> skipping")
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": "empty"})
                    continue
                # Retry transient failures (nav error / timeout / price not
                # rendered). A redirect or a genuine "Currently unavailable" page
                # is a FINAL answer and is NOT retried.
                safe_name = sanitize_filename(url)[:120]
                output_file = os.path.join(output_dir, f"amazon_{idx}_{safe_name}.html")
                result = None
                for attempt in range(1, MAX_ATTEMPTS + 1):
                    page = None
                    try:
                        page = await context.new_page()
                        print(f"\n[Amazon {idx}/{len(urls)}] (attempt {attempt}/{MAX_ATTEMPTS}) Navigating to {url} ...")
                        try:
                            # 30s hard cap so a slow page can't stall the whole run.
                            await page.goto(url, wait_until="load", timeout=30000)
                        except TimeoutError:
                            print(f"⚠️ navigation timeout for {url} after 30s. Continuing anyway...")
                            try:
                                await page.wait_for_load_state("domcontentloaded", timeout=7000)
                            except TimeoutError:
                                pass
                        await asyncio.sleep(10)  # Extra wait to ensure dynamic content loads

                        # Wait randomly for page content to settle
                        await human_delay(3, 6)

                        # 🖱️ Simulate random human-like mouse movement
                        for _ in range(3):
                            x = random.randint(200, 800)
                            y = random.randint(200, 600)
                            await page.mouse.move(x, y, steps=random.randint(5, 15))
                            await human_delay(0.3, 1.5)

                        # 🖱️ Random scrolling
                        for _ in range(2):
                            scroll_y = random.randint(400, 1000)
                            await page.mouse.wheel(0, scroll_y)
                            await human_delay(1, 3)

                        # Extract HTML (resilient to any mid-load client-side navigation)
                        html_content = await get_page_content_safe(page)
                        with open(output_file, "w", encoding="utf-8") as f:
                            f.write(html_content)
                        print(f"✅ HTML saved to {output_file}")

                        # parse (updated: redirect-aware, scoped price)
                        price, redirected = parse_amazon_html(output_file, expected_url=url)

                        if redirected:
                            result = {"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "redirect"}
                            break  # final: different product
                        if price:
                            # Amazon no longer exposes the SM- model on the page; the
                            # writer fills SKU_Amazon from the known per-slot SM code.
                            result = {"url": url, "file": output_file, "price": price, "model": None, "status": "ok"}
                            break  # final: got a price
                        # No price. If the page explicitly says unavailable, that's
                        # a genuine result -> final. Otherwise the price element just
                        # didn't render -> transient -> retry.
                        result = {"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "no_price"}
                        if _looks_unavailable(html_content, "amazon"):
                            print("ℹ️ page marked 'Currently unavailable' -> final, not retrying")
                            break
                        if _amazon_buybox_is_used(html_content):
                            # Used-buybox: the new price is genuinely not offered.
                            # Retrying won't change the condition -> final.
                            print("ℹ️ buybox is a USED offer -> new price not available, not retrying")
                            result["status"] = "used_offer"
                            break
                        print(f"⚠️ price not found & page not marked unavailable (attempt {attempt}/{MAX_ATTEMPTS})")
                    except Exception as e:
                        print(f"❌ Error processing URL {url} (attempt {attempt}/{MAX_ATTEMPTS}): {e}")
                        result = {"url": url, "file": None, "price": None, "model": None, "status": f"error: {e}"}
                    finally:
                        if page:
                            try:
                                await page.close()
                            except Exception:
                                pass
                    # reached only when the attempt was transient (no break)
                    if attempt < MAX_ATTEMPTS:
                        print(f"🔁 retrying in {RETRY_BACKOFF_SEC}s ...")
                        await asyncio.sleep(RETRY_BACKOFF_SEC)

                results.append(result)

            # Save cookies/session state after all pages are processed
            storage_state = await context.storage_state()
            with open(cookies_file, "w", encoding="utf-8") as f:
                json.dump(storage_state, f, ensure_ascii=False, indent=4)
            print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

def parse_amazon_html(html_file_path="amazon.html", expected_url=None):
    """Return (price_text_or_None, redirected_bool).

    - Redirect: compare requested ASIN vs the page's canonical link.
    - Price: scoped to the MAIN price container (corePriceDisplay / priceToPay)
      so we don't pick up sponsored/related prices elsewhere on the page.
    """
    if not os.path.exists(html_file_path):
        print(f"Error: HTML file '{html_file_path}' not found.")
        return None, False

    with open(html_file_path, "r", encoding="utf-8", errors="ignore") as file:
        html_content = file.read()

    # -------- REDIRECT DETECTION --------
    expected_asin = amazon_id_from_url(expected_url) if expected_url else None
    canonical = get_canonical_href(html_content)
    if expected_asin and canonical:
        can_asin = amazon_id_from_url(canonical)
        if can_asin and can_asin != expected_asin:
            print(f"[REDIRECT] requested {expected_asin} but page is {can_asin} -> not available")
            return None, True

    soup = BeautifulSoup(html_content, "lxml")

    # -------- PRICE EXTRACTION (scoped to the main buybox price container) --------
    price = None
    core = (soup.find(id="corePriceDisplay_desktop_feature_div")
            or soup.find(id="corePrice_feature_div")
            or soup.find(id="apex_desktop"))
    if core:
        pt = (core.find(class_="priceToPay")
              or core.find(class_="apexPriceToPay")
              or core)
        price_whole = pt.find("span", {"class": "a-price-whole"})
        price_fraction = pt.find("span", {"class": "a-price-fraction"})
        if price_whole:
            whole = re.sub(r"[^\d,]", "", price_whole.get_text())
            frac = re.sub(r"[^\d]", "", price_fraction.get_text()) if price_fraction else "00"
            price = f"{whole}.{frac or '00'}"
        else:
            for off in pt.find_all("span", {"class": "a-offscreen"}):
                t = off.get_text(strip=True)
                if t:
                    price = t
                    break

    # -------- USED-OFFER GUARD --------
    # If the featured buybox is a USED/renewed device, the price we just read is
    # the USED price. We only track NEW-device prices, so discard it.
    if price and _amazon_buybox_is_used(html_content):
        print(f"⚠️ buybox is a USED offer (price {price}) -> discarding, new price not available")
        price = None

    if price:
        print(f"The price of the product is: {price}")
    else:
        print("Price not found in the HTML file.")

    return price, False


# -----------------------
# BESTBUY-specific logic
# -----------------------
def parse_bestbuy_html(input_file="bestbuy.html", expected_url=None):

    # Load HTML file
    if not os.path.exists(input_file):
        print(f"Error: HTML file '{input_file}' not found.")
        return None, None, False

    with open(input_file, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()

    # -------- REDIRECT DETECTION --------
    expected_code = bestbuy_id_from_url(expected_url) if expected_url else None
    canonical = get_canonical_href(html)
    if expected_code and canonical:
        can_code = bestbuy_id_from_url(canonical)
        if can_code and can_code != expected_code:
            print(f"[REDIRECT] requested {expected_code} but page is {can_code} -> not available")
            return None, None, True

    # -------- PRICE + MODEL from JSON-LD (single, reliable source) --------
    price = None
    model_number = None
    for it in iter_ldjson(html):
        if not isinstance(it, dict):
            continue
        if price is None:
            off = it.get("offers")
            if isinstance(off, dict) and off.get("price"):
                price = str(off["price"])
            elif isinstance(off, list):
                for o in off:
                    if isinstance(o, dict) and o.get("price"):
                        price = str(o["price"]); break
        # Model Number is exposed as a PropertyValue in additionalProperty
        for prop in it.get("additionalProperty", []) or []:
            if isinstance(prop, dict) and str(prop.get("name", "")).lower() == "model number":
                model_number = prop.get("value")
    # regex fallback for model number if not in a parsed object
    if not model_number:
        m = re.search(r'"name"\s*:\s*"Model Number"\s*,\s*"value"\s*:\s*"([^"]+)"', html)
        if m:
            model_number = m.group(1)

    # -------- PRICE fallback: main visible price element (scoped) --------
    if not price:
        soup = BeautifulSoup(html, "lxml")
        el = soup.select_one(
            "div[data-testid='customer-price'] span, "
            "span.font-sans.text-default.text-style-body-md-400.font-500.text-7.leading-7"
        )
        if el:
            price = el.get_text(strip=True)

    print("\n--- Extracted Product Data (BestBuy) ---")
    print(f"File: {input_file}")
    print(f"Price: {price}")
    print(f"Model Number: {model_number}")
    print("--------------------------------\n")

    return price, model_number, False

def _find_free_port():
    """Return an OS-assigned free TCP port on localhost."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


def _find_real_chrome():
    """Locate the REAL Google Chrome binary (NOT Playwright's bundled Chromium).

    BestBuy's bot protection is confirmed to pass with real Chrome; bundled
    Chromium differs in fingerprint (UA-brand "Chromium", no Widevine/H.264,
    different userAgentData) and may be flagged. So for BestBuy we insist on
    real Chrome.

    Resolution order:
      1. $CHROME_BIN / $CHROME_PATH env var (set this on EC2 if Chrome is in a
         non-standard location).
      2. `chrome`/`google-chrome`/`google-chrome-stable` on PATH (Linux; the
         GitHub Actions ubuntu runner ships google-chrome-stable).
      3. Standard Windows install locations (for local Windows 11 runs).
    Returns the path, or None if not found.
    """
    env = os.environ.get("CHROME_BIN") or os.environ.get("CHROME_PATH")
    if env and os.path.exists(env):
        return env

    for name in ("google-chrome-stable", "google-chrome", "chrome",
                 "chromium-browser", "chromium"):
        found = shutil.which(name)
        if found:
            return found

    for candidate in (
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
        "/opt/google/chrome/chrome",
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    ):
        if candidate and os.path.exists(candidate):
            return candidate

    return None


async def _launch_chrome_cdp(playwright, extra_args=None):
    """Launch a REAL Chrome/Chromium process with a remote-debugging port and
    connect Playwright to it over CDP.

    Why CDP instead of playwright.chromium.launch() for BestBuy: launch() starts
    Chromium with Playwright's automation switches (e.g. --enable-automation,
    AutomationControlled), which BestBuy's bot protection fingerprints. By
    starting Chrome ourselves with only the flags we choose and attaching over
    the DevTools Protocol, the browser looks like an ordinary Chrome instance.

    Each call launches a fresh process with a fresh throwaway --user-data-dir, so
    every attempt gets a brand-new HTTP/2 connection AND a pristine, cookieless
    profile (the two things that previously broke URL 2+).

    Returns (browser, proc, profile_dir). Caller must close browser, kill proc,
    and remove profile_dir.
    """
    port = _find_free_port()
    profile_dir = tempfile.mkdtemp(prefix="bb_cdp_profile_")
    # Use the REAL Google Chrome binary (confirmed to pass BestBuy). Do NOT fall
    # back to bundled Chromium — its fingerprint differs and may be flagged.
    chrome_path = _find_real_chrome()
    if not chrome_path:
        shutil.rmtree(profile_dir, ignore_errors=True)
        raise RuntimeError(
            "Real Google Chrome not found. Install it (Actions ubuntu runner "
            "ships google-chrome-stable; on EC2 install google-chrome-stable) "
            "or set the CHROME_BIN env var to its path.")

    args = [
        chrome_path,
        f"--remote-debugging-port={port}",
        f"--user-data-dir={profile_dir}",
        "--no-first-run",
        "--no-default-browser-check",
        "--no-sandbox",
        "--disable-setuid-sandbox",
        "--disable-dev-shm-usage",
        "--disable-gpu",
        "about:blank",
    ]
    if extra_args:
        args.extend(extra_args)

    # Headful: BestBuy won't serve a headless browser. On a display-less server
    # (EC2 / GitHub Actions) this runs under Xvfb, which supplies $DISPLAY.
    proc = subprocess.Popen(
        args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    # Wait for the CDP endpoint to come up, then read the WebSocket URL.
    cdp_http = f"http://127.0.0.1:{port}"
    ws_url = None
    for _ in range(60):  # up to ~30s
        try:
            with urllib.request.urlopen(f"{cdp_http}/json/version", timeout=1) as r:
                ws_url = json.loads(r.read().decode()).get("webSocketDebuggerUrl")
            if ws_url:
                break
        except Exception:
            await asyncio.sleep(0.5)
    if not ws_url:
        # cleanup before raising
        try:
            proc.kill()
        except Exception:
            pass
        shutil.rmtree(profile_dir, ignore_errors=True)
        raise RuntimeError("Chrome CDP endpoint did not come up in time")

    browser = await playwright.chromium.connect_over_cdp(ws_url)
    return browser, proc, profile_dir


async def save_bestbuy_htmls(
    urls,
    output_dir="outputs",
    cookies_file="bestbuy_cookies.json",
    headless=True,
):
    """
    Loop over list of BestBuy URLs, save each page's HTML to output_dir,
    parse with parse_bestbuy_html and return results list.
    """
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        # ROOT CAUSE of the "only the first URL works" failure: a single browser
        # (and single context) was reused for every URL. Chromium keeps the HTTP/2
        # connection to bestbuy.com ALIVE and reuses it across pages. BestBuy's bot
        # protection flags that connection after the first request and then RESETS
        # every subsequent stream on it -> net::ERR_HTTP2_PROTOCOL_ERROR on URL 2+.
        # A new page on the same context reuses the same poisoned connection, so it
        # never recovers.
        #
        # FIX: launch a COMPLETELY FRESH browser per attempt (guarantees a brand-new
        # HTTP/2 connection every time), persisting cookies to disk between browsers
        # so the session still carries over. This is the strongest form of the
        # "fresh context per URL" trick the reference scraper uses.
        #
        # Browser is launched per-attempt via CDP (see _launch_chrome_cdp): we
        # start a REAL Chrome process ourselves and attach over the DevTools
        # Protocol, instead of p.chromium.launch() which stamps automation flags
        # BestBuy fingerprints. Each attempt = fresh process + fresh throwaway
        # profile = fresh HTTP/2 connection + pristine cookieless session.
        #
        # Other BestBuy specifics (unchanged):
        #   - Headful: BestBuy won't serve a headless browser. On a display-less
        #     server (EC2 / GitHub Actions) run under Xvfb.
        #   - Default HTTP/2 (no --disable-http2, which returned empty shells).
        #   - No custom user_agent (a spoofed UA on real Linux Chromium is a tell).
        #   - Block image/media/font requests: fewer HTTP/2 streams to reset.

        results = []
        for idx, url in enumerate(urls, start=1):
            # empty slot (e.g. product not yet listed): keep the position so
            # results stay aligned with the product groups, but skip cleanly.
            if not url or not url.strip():
                print(f"\n[BestBuy {idx}/{len(urls)}] empty URL slot -> skipping")
                results.append({"url": url, "file": None, "price": None, "model": None, "status": "empty"})
                continue
            # Retry transient failures (nav error / timeout / JSON-LD not
            # rendered). A redirect or a genuine sold-out page is final.
            safe_name = sanitize_filename(url)[:120]
            output_file = os.path.join(output_dir, f"bestbuy_{idx}_{safe_name}.html")
            result = None
            for attempt in range(1, MAX_ATTEMPTS + 1):
                browser = None
                proc = None
                profile_dir = None
                context = None
                page = None
                try:
                    # FRESH Chrome via CDP -> fresh HTTP/2 connection + a pristine,
                    # throwaway profile for this URL/attempt. COOKIELESS by design:
                    # BestBuy plants a bot-detection/session token in its response
                    # cookies on the first visit; replaying it poisoned URL 2+
                    # (net::ERR_HTTP2_PROTOCOL_ERROR). A brand-new --user-data-dir
                    # each time guarantees no cookie/session carries over.
                    browser, proc, profile_dir = await _launch_chrome_cdp(p)

                    # Over CDP the fresh Chrome already exposes a default context
                    # (browser.contexts[0]); since the profile is brand-new it is
                    # pristine and cookieless, so we use it directly.
                    context = (browser.contexts[0] if browser.contexts
                               else await browser.new_context())

                    page = await context.new_page()

                    # Apply anti-bot stealth patches (BestBuy only). Masks the
                    # automation fingerprints (navigator.webdriver, headless
                    # hints, etc.) that BestBuy's protection checks. No-op if the
                    # playwright-stealth package isn't installed. Must run before
                    # navigation so the patches are in place when page scripts run.
                    try:
                        await _apply_stealth(page)
                    except Exception as _se:
                        print(f"⚠️ stealth patch failed (continuing without it): {_se}")

                    # Abort image/media/font requests. Each is an HTTP/2 stream on
                    # the connection; blocking them leaves only the document +
                    # JSON/JS we actually parse (lighter, faster, fewer resets).
                    async def _block_heavy(route, request):
                        if request.resource_type in ("image", "media", "font"):
                            await route.abort()
                        else:
                            await route.continue_()
                    await page.route("**/*", _block_heavy)

                    print(f"\n[BestBuy {idx}/{len(urls)}] (attempt {attempt}/{MAX_ATTEMPTS}) Navigating to {url} ...")

                    try:
                        # wait_until="domcontentloaded" (NOT "load"): returning at
                        # DOMContentLoaded avoids waiting on the flaky subresource
                        # streams; the JS-injected price populates during the wait
                        # below. Generous 180s cap matches the reference.
                        await page.goto(url, wait_until="domcontentloaded", timeout=180000)
                        await asyncio.sleep(10)  # extra wait to ensure stability
                    except TimeoutError:
                        print(f"⚠️ navigation timeout for {url} after 180s. Continuing anyway...")

                    # Wait randomly for page content to settle
                    await human_delay(3, 6)

                    # 🖱️ Simulate random human-like mouse movement
                    for _ in range(3):
                        x = random.randint(200, 800)
                        y = random.randint(200, 600)
                        await page.mouse.move(x, y, steps=random.randint(5, 15))
                        await human_delay(0.3, 1.5)

                    # 🖱️ Random scrolling
                    for _ in range(2):
                        scroll_y = random.randint(400, 1000)
                        await page.mouse.wheel(0, scroll_y)
                        await human_delay(1, 3)

                    # The price/model come from a JSON-LD block that BestBuy
                    # injects via client-side JS. Wait for that JSON-LD (with an
                    # "offers" field) before capturing, so we don't save a
                    # half-loaded page that parses to Price=None / Model=None.
                    try:
                        await page.wait_for_function(
                            """() => {
                                const s = document.querySelectorAll('script[type="application/ld+json"]');
                                for (const el of s) {
                                    if (el.textContent && el.textContent.indexOf('"offers"') !== -1) return true;
                                }
                                return false;
                            }""",
                            timeout=20000,
                        )
                    except Exception:
                        print("⚠️ BestBuy JSON-LD offers not detected within 20s; saving page anyway")

                    # Extract HTML (resilient to BestBuy's mid-load client-side
                    # navigation which used to make page.content() throw)
                    html_content = await get_page_content_safe(page)
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    print(f"✅ HTML saved to {output_file}")

                    # NOTE: intentionally do NOT save cookies for BestBuy. Persisting
                    # BestBuy's session token and replaying it is what poisoned URL 2+
                    # (see the cookieless context comment above). Every page stays
                    # pristine, matching the known-good reference scraper.

                    # Parse (updated: redirect-aware, JSON-LD)
                    price, model, redirected = parse_bestbuy_html(output_file, expected_url=url)

                    if redirected:
                        result = {"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "redirect"}
                        break  # final: different product
                    if price:
                        result = {"url": url, "file": output_file, "price": price, "model": model, "status": "ok"}
                        break  # final: got a price
                    result = {"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "no_price"}
                    if _looks_unavailable(html_content, "bestbuy"):
                        print("ℹ️ page marked sold out / no longer available -> final, not retrying")
                        break
                    print(f"⚠️ price not found & page not marked unavailable (attempt {attempt}/{MAX_ATTEMPTS})")
                except Exception as e:
                    print(f"❌ Error processing URL {url} (attempt {attempt}/{MAX_ATTEMPTS}): {e}")
                    result = {"url": url, "file": None, "price": None, "model": None, "status": f"error: {e}"}
                finally:
                    # Tear down EVERYTHING so the next attempt/URL starts on a
                    # brand-new Chrome process + connection + profile.
                    for closable in (page, context, browser):
                        if closable:
                            try:
                                await closable.close()
                            except Exception:
                                pass
                    # browser.close() only DISCONNECTS the CDP session; the real
                    # Chrome process we spawned must be killed explicitly, and its
                    # throwaway profile dir removed, or they'd leak every attempt.
                    if proc:
                        try:
                            proc.kill()
                            proc.wait(timeout=10)
                        except Exception:
                            pass
                    if profile_dir:
                        shutil.rmtree(profile_dir, ignore_errors=True)
                if attempt < MAX_ATTEMPTS:
                    print(f"🔁 retrying in {RETRY_BACKOFF_SEC}s ...")
                    await asyncio.sleep(RETRY_BACKOFF_SEC)

            results.append(result)

    return results

# -----------------------
# SAMSUNG-specific logic
# -----------------------
async def wait_network_idle(page, timeout=15000):
    """Wait until network becomes idle (0 active requests)."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
        await asyncio.sleep(10)  # extra wait to ensure stability | not sure if this is the right approach
    except TimeoutError:
        print("⚠️ networkidle timeout — continuing anyway")

def extract_sku_from_url(url: str):
    """Extract SKU value from the given URL (looks for 'sku-<value>' or 'sm-<value>')."""
    if not url:
        return None
    
    # Try to find sku- first
    m = re.search(r"sku-([A-Za-z0-9-]+)", url, re.IGNORECASE)
    if m:
        return m.group(1)
    
    # If not found, try to find sm-
    m = re.search(r"(sm-[A-Za-z0-9-]+)", url, re.IGNORECASE)
    if m:
        return m.group(1)
    
    return None


def extract_price(filename, expected_url=None):
    """Return (price_text_or_None, redirected_bool) for a saved Samsung page.

    Updated for the current UI: the old #device_info aria-checked radios are gone.
    The reliable source is the JSON-LD Product offer, keyed to the exact SKU, so
    we never pick up a sibling variant's price. Redirects are detected via the
    page's canonical link.
    """
    if not os.path.exists(filename):
        print(f"❌ File not found for parsing: {filename}")
        return None, False

    with open(filename, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()

    expected_sku = extract_sku_from_url(expected_url) if expected_url else None
    expected_sku = expected_sku.lower() if expected_sku else None

    # -------- REDIRECT DETECTION --------
    canonical = get_canonical_href(html)
    if expected_sku and canonical:
        can_sku = extract_sku_from_url(canonical)
        can_sku = can_sku.lower() if can_sku else None
        if can_sku and can_sku != expected_sku:
            print(f"[REDIRECT] requested {expected_sku} but page is {can_sku} -> not available")
            return None, True

    # -------- PRICE from JSON-LD offer keyed to the SKU --------
    price = None
    for it in iter_ldjson(html):
        if isinstance(it, dict) and it.get("sku"):
            if expected_sku and str(it["sku"]).lower() != expected_sku:
                continue
            off = it.get("offers")
            if isinstance(off, dict) and off.get("price"):
                price = str(off["price"]); break

    print("🔎 Extracted Price:", price)
    return price, False

async def save_samsung_htmls(
    urls,
    output_dir="outputs",
    cookies_file="samsung_cookies.json",
    headless=True,
):
    """
    Loop over list of Samsung product URLs, save each page's HTML to output_dir,
    parse price and sku using the same logic you provided, and return results list.
    """
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)   # keep visible by default per original
        # Create or reuse context
        # if os.path.exists(cookies_file):
        #     print("🍪 Loading existing cookies/session...")
        #     context = await browser.new_context(storage_state=cookies_file)
        # else:
        print("🆕 No cookies found, creating a new session...")

        results = []
        try:
            for idx, url in enumerate(urls, start=1):
                # empty slot (e.g. product not yet listed): keep the position so
                # results stay aligned with the product groups, but skip cleanly.
                if not url or not url.strip():
                    print(f"\n[Samsung {idx}/{len(urls)}] empty URL slot -> skipping")
                    results.append({"url": url, "file": None, "price": None, "sku": None, "status": "empty"})
                    continue
                safe_name = sanitize_filename(url)
                output_file = os.path.join(output_dir, f"samsung_{idx}_{safe_name}.html")
                sku = extract_sku_from_url(url)

                # Retry transient failures (nav error / timeout / #device_info not
                # loaded / price not rendered). A redirect or a genuine sold-out /
                # coming-soon page is final and is NOT retried.
                result = None
                for attempt in range(1, MAX_ATTEMPTS + 1):
                    context = None
                    page = None
                    try:
                        # fresh context per attempt (isolates cookies/storage)
                        context = await browser.new_context(
                            user_agent=(
                                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                "AppleWebKit/537.36 (KHTML, like Gecko) "
                                "Chrome/120.0.0.0 Safari/537.36"
                            ),
                            viewport={"width": 1600, "height": 900},
                        )

                        page = await context.new_page()
                        print(f"\n[Samsung {idx}/{len(urls)}] (attempt {attempt}/{MAX_ATTEMPTS}) Navigating to {url} ...")
                        try:
                            # 30s hard cap so a slow page can't stall the whole run.
                            await page.goto(url, wait_until="load", timeout=30000)
                        except TimeoutError:
                            print(f"⚠️ navigation timeout for {url} after 30s. Continuing anyway...")

                        print("Waiting for network to be idle...")
                        await wait_network_idle(page, timeout=20000)

                        print("Waiting for #device_info box...")
                        device_info_ok = True
                        try:
                            await page.wait_for_selector("#device_info", timeout=20000)
                            # Extra wait for prices inside #device_info
                            await page.wait_for_selector("#device_info span", timeout=15000)
                        except TimeoutError:
                            print("❌ #device_info did NOT load — Samsung blocked or loaded too slowly.")
                            device_info_ok = False

                        # Save HTML (resilient to any mid-load client-side navigation)
                        html = await get_page_content_safe(page)
                        with open(output_file, "w", encoding="utf-8") as f:
                            f.write(html)
                        print(f"✅ HTML saved to {output_file}")

                        if not device_info_ok:
                            # transient (page structure never appeared) -> retry
                            result = {"url": url, "file": output_file, "price": None, "sku": sku, "status": "partial: no device_info"}
                            print(f"⚠️ #device_info missing (attempt {attempt}/{MAX_ATTEMPTS})")
                        else:
                            # Parse saved HTML (updated: redirect-aware, JSON-LD by SKU)
                            price, redirected = extract_price(output_file, expected_url=url)
                            if redirected:
                                print("[REDIRECT] Samsung redirect -> not available")
                                result = {"url": url, "file": output_file, "price": NOT_AVAILABLE, "sku": NOT_AVAILABLE, "status": "redirect"}
                                break  # final: different product
                            elif price:
                                print("🔎 Final extracted values — Price:", price, "SKU:", sku)
                                result = {"url": url, "file": output_file, "price": price, "sku": sku, "status": "ok"}
                                break  # final: got a price
                            else:
                                result = {"url": url, "file": output_file, "price": NOT_AVAILABLE, "sku": NOT_AVAILABLE, "status": "no_price"}
                                if _looks_unavailable(html, "samsung"):
                                    print("ℹ️ page marked sold out / coming soon -> final, not retrying")
                                    break
                                print(f"⚠️ price not found & page not marked unavailable (attempt {attempt}/{MAX_ATTEMPTS})")

                        # tiny cooperative yield
                        await human_delay_short()
                    except Exception as e:
                        print(f"❌ Error processing URL {url} (attempt {attempt}/{MAX_ATTEMPTS}): {e}")
                        result = {"url": url, "file": None, "price": None, "sku": None, "status": f"error: {e}"}
                    finally:
                        try:
                            if page:
                                await page.close()
                        except Exception:
                            pass
                        try:
                            if context:
                                await context.close()
                        except Exception:
                            pass
                    if attempt < MAX_ATTEMPTS:
                        print(f"🔁 retrying in {RETRY_BACKOFF_SEC}s ...")
                        await asyncio.sleep(RETRY_BACKOFF_SEC)

                results.append(result)

            # Write cookies/session state once more at the end
            # storage = await context.storage_state()
            # with open(cookies_file, "w", encoding="utf-8") as f:
            #     json.dump(storage, f, indent=2)
            # print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

# -----------------------
# Combined main
# -----------------------
async def main():
    # Capture the run's START time in EST. We snap THIS (not the finish time) to
    # the nearest scheduled 6-hour mark, so the Timestamp column reflects the
    # slot the run was launched for and is unaffected by how long scraping takes.
    run_start_est = datetime.datetime.now(datetime.timezone.utc).astimezone(
        ZoneInfo("US/Eastern"))

    # Replace/extend these lists with the product URLs you want to iterate over
    amazon_urls = [
    #watches


    #Galaxy Watch9 ( 40 mm) cream bluetooth
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BHP9XW/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",
    #Galaxy Watch9 ( 40 mm) cream lte
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BF1BFS/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",

    #Galaxy Watch9 ( 40 mm) graphite bluetooth
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BPZXZC/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",
    #Galaxy Watch9 ( 40 mm) graphite lte
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BL3119/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",

    #Galaxy Watch9 (Bluetooth, 44 mm) sliver bluetooth
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BH59VZ/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",
    #Galaxy Watch9 ( 44 mm) sliver lte
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9B66J2R/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",

    #Galaxy Watch9 ( 44 mm) graphite bluetooth
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BKX9S9/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",
    #Galaxy Watch9 (44 mm) graphite lte
    "https://www.amazon.com/Samsung-Galaxy-Graphite-Bluetooth-Smartwatch/dp/B0H9BKV7J1/ref=sr_1_1?crid=2BWX5OHPW4EF2&dib=eyJ2IjoiMSJ9.P5DKyv57_CHmoZ2U9NKPJLP4FpJ3vH_7jbbHaULy7AfJRCd1LrJ2DswEZjLLpYka9bEThlNcxaPP8LTFSre-8pnZR-B0Pzdufqbdgg8IfPeLoMAaR339mJtSoM1N24soAJ_cayN_bAVoKPFy-2ipfEGJkgVUEUZrE0nU6LWDF46PvZVHYAP9JOPOkWKd6gurF5GIi8XXTRL_eEI9dQQRjITnO15mroIP5sadnS-TvtELoiIsSBJG0oCGZDeSdZTgHsEpS6OR2hVSHjcFnBKADLox3KuQrmZIL8PJ7BlGyHg.GieqM3OkERmL2_NRRpK0OB_ofnfjWmAX0zaE9NPRemI&dib_tag=se&keywords=Galaxy%2BWatch9&qid=1784789591&s=electronics&sprefix=%2Celectronics%2C299&sr=1-1&th=1",


    #Samsung Galaxy Watch Ultra2 , Titanium , 47mm LTE Color: Titanium gray
    "https://www.amazon.com/Samsung-Galaxy-Ultra2-Titanium-Smartwatch/dp/B0H6NRPGT7?ref_=ast_sto_dp&th=1",

    #Samsung Galaxy Watch Ultra2 ,  47mm LTE Color: Titanium Silver
    "https://www.amazon.com/Samsung-Galaxy-Ultra2-Titanium-Smartwatch/dp/B0H6N8SCW5?ref_=ast_sto_dp&th=1"
    
    ]

    bestbuy_urls = [

    ]

    samsung_urls = [
    #Galaxy Watch9 (Bluetooth, 40 mm) cream bluetooth
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-40mm-cream-bluetooth-sku-sm-l340nzeaxaa/",
    #Galaxy Watch9 (40 mm) cream lte
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-40mm-cream-lte-sku-sm-l345uzedxaa/",

    #Galaxy Watch9 (Bluetooth, 40 mm) graphite bluetooth
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-40mm-graphite-bluetooth-sku-sm-l340nzkaxaa/",
    #Galaxy Watch9 ( 40 mm) graphite lte
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-40mm-graphite-lte-sku-sm-l345uzkaxaa/",

    #Galaxy Watch9 (Bluetooth, 44 mm) sliver bluetooth
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-44mm-silver-bluetooth-sku-sm-l350nzsaxaa/",
    #Galaxy Watch9 ( 44 mm) sliver lte
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-44mm-silver-lte-sku-sm-l355uzsdxaa/",

    #Galaxy Watch9 (Bluetooth, 44 mm) graphite bluetooth
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-44mm-graphite-bluetooth-sku-sm-l350nzkaxaa/",
    #Galaxy Watch9 ( 44 mm) graphite lte
    "https://www.samsung.com/us/watches/galaxy-watch9/buy/galaxy-watch9-44mm-graphite-lte-sku-sm-l355uzkdxaa/",





    #Galaxy Watch Ultra2 47mm lte titanium gray , band - olive
    "https://www.samsung.com/us/watches/galaxy-watch-ultra2/buy/galaxy-watch-ultra2-47mm-titanium-gray-lte-sku-sm-l715uzkaxaa/",
    #Galaxy Watch Ultra2 47mm lte titanium silver, band - olive
    "https://www.samsung.com/us/watches/galaxy-watch-ultra2/buy/galaxy-watch-ultra2-47mm-titanium-silver-lte-sku-sm-l715uzsaxaa/",

    ]

    print("\n=== Running Amazon scraper ===")
    am_res = await save_amazon_htmls(amazon_urls, output_dir="outputs", cookies_file="amazon_cookies.json", headless=True)
    print("\nAmazon Summary:")
    for r in am_res:
        print(r)

    print("\n=== Running BestBuy scraper ===")
    bb_res = await save_bestbuy_htmls(bestbuy_urls, output_dir="outputs", cookies_file="bestbuy_cookies.json", headless=True)
    print("\nBestBuy Summary:")
    for r in bb_res:
        print(r)

    print("\n=== Running Samsung scraper ===")
    sam_res = await save_samsung_htmls(samsung_urls, output_dir="outputs", cookies_file="samsung_cookies.json", headless=True)
    print("\nSamsung Summary:")
    for r in sam_res:
        print(r)

    # -----------------------
    # Write results in the SAME layout as Price Comparisons_v3_WIP.xlsx
    # (one row per run; prices + SKU columns only; 'vs' formulas left blank).
    # am_res / bb_res / sam_res are in URL order, i.e. slot order, so index s
    # maps directly to product group s.
    # -----------------------
    # The run is scheduled 4x/day to launch on 00:00 / 06:00 / 12:00 / 18:00 EST,
    # but cron / startup jitter means run_start_est is a few minutes off. Snap the
    # START time (captured at the top of main(), NOT this finish time) to the
    # nearest 6-hour mark so the Timestamp column always shows one of the four
    # exact scheduled times, independent of how long scraping took.
    # Rounding to the nearest multiple of 6h naturally yields 0/6/12/18, and rolls
    # over to the next day's 00:00 when the run launches just before midnight.
    _mins = run_start_est.hour * 60 + run_start_est.minute + run_start_est.second / 60
    _snapped = round(_mins / 360) * 360          # nearest 6h (360 min) boundary
    est_snapped = run_start_est.replace(hour=0, minute=0, second=0, microsecond=0) \
        + datetime.timedelta(minutes=_snapped)   # +1440 rolls into the next day
    ts_str = est_snapped.strftime("%d %b %Y, %H:%M")   # e.g. "05 Dec 2025, 06:00"

    excel_file = os.path.join("outputs", "results.xlsx")
    save_results_wip_format(am_res, bb_res, sam_res, samsung_urls, ts_str, excel_file)

if __name__ == "__main__":
    asyncio.run(main())
    file_path = "outputs/results.xlsx"

    # column_references = [
    #     "a","d","ar","x","e","as","y","blank column",
    #     "i","aw","ac","j","ax","ad","blank column",
    #     "n","bb","ah","o","bc","ai","blank column",
    #     "s","bg","am","t","bh","an"
    # ]

    # created = copy_columns_by_references(
    #     file_path=file_path,
    #     column_refs=column_references,
    #     source_sheet_name=None,      # None => use first sheet; or set "Sheet1"
    #     new_sheet_base_name="SelectedColumns"
    # )
    # print(f"Created sheet: {created} in {file_path}")
